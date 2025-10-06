"""
Gestionnaire de résultats de jeux pour le bot Telegram
Stocke les parties où le premier groupe a exactement 3 cartes différentes
"""
import re
import yaml
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class GameResultsManager:
    """Gestionnaire pour stocker les résultats des jeux de cartes"""
    
    def __init__(self):
        # Répertoire pour stocker les données
        self.data_dir = Path("data")
        self.data_dir.mkdir(exist_ok=True)
        
        # Fichier de données des résultats
        self.results_file = self.data_dir / "game_results.yaml"
        
        # Initialiser le fichier s'il n'existe pas
        if not self.results_file.exists():
            self._save_yaml([])
        
        print("✅ Gestionnaire de résultats initialisé")
    
    def _load_yaml(self) -> List[Dict[str, Any]]:
        """Charge les résultats depuis le fichier YAML"""
        try:
            if self.results_file.exists():
                with open(self.results_file, 'r', encoding='utf-8') as f:
                    data = yaml.safe_load(f)
                    return data if isinstance(data, list) else []
            return []
        except Exception as e:
            print(f"❌ Erreur chargement résultats: {e}")
            return []
    
    def _save_yaml(self, data: List[Dict[str, Any]]):
        """Sauvegarde les résultats dans le fichier YAML"""
        try:
            with open(self.results_file, 'w', encoding='utf-8') as f:
                yaml.dump(data, f, allow_unicode=True, default_flow_style=False, indent=2)
        except Exception as e:
            print(f"❌ Erreur sauvegarde résultats: {e}")
    
    def extract_game_number(self, message: str) -> Optional[int]:
        """Extrait le numéro de jeu du message"""
        try:
            # Chercher les patterns comme "#N 123", "#N123", "#N60.", etc.
            match = re.search(r"#N\s*(\d+)\.?", message, re.IGNORECASE)
            if match:
                return int(match.group(1))
            
            # Pattern alternatif
            match = re.search(r"jeu\s*#?\s*(\d+)", message, re.IGNORECASE)
            if match:
                return int(match.group(1))
            
            return None
        except Exception as e:
            print(f"❌ Erreur extraction numéro: {e}")
            return None
    
    def extract_parentheses_groups(self, message: str) -> List[str]:
        """Extrait le contenu des parenthèses du message"""
        try:
            return re.findall(r"\(([^)]*)\)", message)
        except Exception:
            return []
    
    def count_cards(self, group_str: str) -> int:
        """Compte le nombre de symboles de cartes dans un groupe"""
        emoji_symbols = ['♠️', '♥️', '♦️', '♣️']
        simple_symbols = ['♠', '♥', '♦', '♣']
        
        # Compter les emojis
        temp_str = group_str
        emoji_count = 0
        for emoji in emoji_symbols:
            count = temp_str.count(emoji)
            emoji_count += count
            temp_str = temp_str.replace(emoji, 'X')
        
        # Compter les symboles simples
        simple_count = 0
        for symbol in simple_symbols:
            simple_count += temp_str.count(symbol)
        
        return emoji_count + simple_count
    
    def has_different_suits(self, group_str: str) -> bool:
        """
        Vérifie si un groupe contient 3 cartes de SYMBOLES DIFFÉRENTS
        Retourne True si les 3 symboles sont tous différents
        
        Combinaisons valides (24 au total) - toutes permutations de:
        ♠️ ❤️ ♣️ | ♠️ ❤️ ♦️ | ♠️ ♣️ ♦️ | ❤️ ♣️ ♦️
        
        Supporte: ♠️ ♠ | ❤️ ❤ ♥️ ♥ | ♦️ ♦ | ♣️ ♣
        """
        # Normaliser TOUS les symboles de cœur vers ♥
        normalized = group_str.replace('❤️', '♥').replace('❤', '♥').replace('♥️', '♥')
        
        # Normaliser les autres symboles (enlever le modificateur emoji U+FE0F)
        normalized = normalized.replace('♠️', '♠').replace('♦️', '♦').replace('♣️', '♣')
        
        # Les 4 symboles de base
        suits = ['♠', '♥', '♦', '♣']
        
        # Compter chaque symbole
        suit_counts = {}
        for suit in suits:
            count = normalized.count(suit)
            if count > 0:
                suit_counts[suit] = count
        
        # Validation stricte pour 3 couleurs différentes:
        # - Exactement 3 symboles distincts présents
        # - Chaque symbole apparaît exactement 1 fois
        if len(suit_counts) != 3:
            return False
        
        return all(count == 1 for count in suit_counts.values())
    
    def determine_winner(self, message: str, first_group: str, second_group: str) -> Optional[str]:
        """
        Détermine le gagnant (Joueur ou Banquier) en fonction du message
        Retourne 'Joueur', 'Banquier' ou None (match nul ou non déterminé)
        """
        message_upper = message.upper()
        
        # NOUVELLE DÉTECTION: Symbole ▶️ indique le gagnant
        # Format: ▶️ X(cartes) signifie que ce groupe a gagné
        parts_split = message.split(' - ')
        if len(parts_split) >= 2:
            first_part = parts_split[0]
            second_part = parts_split[1]
            
            # Vérifier où se trouve le symbole ▶️
            if '▶️' in first_part:
                return 'Joueur'
            elif '▶️' in second_part:
                return 'Banquier'
        
        # Chercher les indicateurs de victoire
        if any(indicator in message_upper for indicator in ['JOUEUR', 'PLAYER', 'J GAGNE', 'VICTOIRE J']):
            return 'Joueur'
        elif any(indicator in message_upper for indicator in ['BANQUIER', 'BANKER', 'B GAGNE', 'VICTOIRE B']):
            return 'Banquier'
        
        # Logique alternative: compter les points ou détecter des patterns
        # Si le message contient des émojis spécifiques
        if '🎯' in message or '✅' in message:
            # Analyse basée sur la position ou le contexte
            parts = message.split('|')
            if len(parts) >= 2:
                first_part = parts[0]
                second_part = parts[1] if len(parts) > 1 else ''
                
                # Si le premier groupe (Joueur) est marqué avec ✅
                if '✅' in first_part or '🎯' in first_part:
                    return 'Joueur'
                elif '✅' in second_part or '🎯' in second_part:
                    return 'Banquier'
            else:
                # Si pas de |, analyser différemment
                parts_split = message.split(' - ')
                if len(parts_split) >= 2:
                    if '✅' in parts_split[0] or '🎯' in parts_split[0]:
                        return 'Joueur'
                    elif '✅' in parts_split[1] or '🎯' in parts_split[1]:
                        return 'Banquier'
        
        # Chercher "P" ou "B" après les parenthèses
        pattern = r'\)\s*-\s*\([^)]*\)\s*([PB])'
        match = re.search(pattern, message, re.IGNORECASE)
        if match:
            winner_letter = match.group(1).upper()
            return 'Joueur' if winner_letter == 'P' else 'Banquier'
        
        # Si aucun gagnant clair, retourner None (match nul)
        return None
    
    def extract_datetime_from_message(self, message: str) -> Tuple[str, str]:
        """Extrait la date et l'heure du message si disponible"""
        try:
            # Chercher le pattern de date (DD/MM/YYYY ou similaire)
            date_match = re.search(r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})', message)
            # Chercher le pattern d'heure (HH:MM ou HH:MM:SS)
            time_match = re.search(r'(\d{1,2}:\d{2}(?::\d{2})?)', message)
            
            if date_match and time_match:
                date_str = date_match.group(1)
                time_str = time_match.group(1)
                
                # Convertir au format standardisé YYYY-MM-DD
                date_parts = re.split(r'[/\-\.]', date_str)
                if len(date_parts) == 3:
                    day, month, year = date_parts
                    if len(year) == 2:
                        year = '20' + year
                    formatted_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    
                    # Ajouter les secondes si manquantes
                    if len(time_str) == 5:  # HH:MM
                        time_str = time_str + ':00'
                    
                    return formatted_date, time_str
        except:
            pass
        
        # Fallback: utiliser l'heure actuelle
        now = datetime.now()
        return now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S')
    
    def process_message(self, message: str) -> Tuple[bool, Optional[str]]:
        """
        Traite un message et stocke le résultat si les conditions sont remplies
        
        NOUVELLES RÈGLES:
        - Ne PAS contenir ⏰ (message en cours)
        - Ne PAS contenir 🔰 (on ignore ces messages)
        - Doit contenir ✅ (message finalisé)
        - Si premier groupe a 3 cartes différentes → Victoire JOUEUR
        - Si deuxième groupe a 3 cartes différentes → Victoire BANQUIER
        - Si les deux ont 3 cartes différentes → NE RIEN enregistrer
        - Ne pas enregistrer les numéros consécutifs (N puis N+1)
        
        Retourne: (succès, message_info)
        """
        try:
            # Log du message complet pour debug
            print(f"📩 Message reçu: {message[:150]}...")
            
            # VÉRIFICATION 1: Le message NE doit PAS être en cours
            if '⏰' in message:
                print(f"⏰ Message en cours d'édition, attente de finalisation...")
                return False, "Message en cours d'édition (symbole ⏰)"
            
            # VÉRIFICATION 2: Le message NE doit PAS contenir 🔰
            if '🔰' in message:
                print(f"🔰 Message avec symbole 🔰, on ignore")
                return False, "Message avec symbole 🔰 (ignoré)"
            
            # VÉRIFICATION 3: Le message doit contenir ✅
            if '✅' not in message:
                print(f"⚠️ Message non finalisé (pas de ✅)")
                return False, "Message non finalisé (pas de symbole ✅)"
            
            print(f"✅ Message finalisé détecté, traitement en cours...")
            
            # Extraire le numéro de jeu
            game_number = self.extract_game_number(message)
            if game_number is None:
                print(f"❌ Pas de numéro de jeu trouvé dans: {message[:100]}")
                return False, "Pas de numéro de jeu trouvé"
            
            # Charger les résultats existants
            results = self._load_yaml()
            
            # Vérifier si ce jeu n'est pas déjà stocké
            if any(r.get('numero') == game_number for r in results):
                print(f"ℹ️ Jeu #{game_number} déjà enregistré")
                return False, f"Jeu #{game_number} déjà enregistré"
            
            # Vérifier les numéros consécutifs contre TOUS les numéros enregistrés
            if results:
                for result in results:
                    stored_number = result.get('numero', 0)
                    if game_number == stored_number + 1:
                        print(f"⚠️ Numéro consécutif détecté (numéro {stored_number} déjà enregistré, actuel: {game_number}), message ignoré")
                        return False, f"Numéro consécutif ignoré ({stored_number} → {game_number})"
            
            # Extraire les groupes de parenthèses
            groups = self.extract_parentheses_groups(message)
            if len(groups) < 2:
                print(f"❌ Pas assez de groupes de parenthèses: {groups}")
                return False, "Pas assez de groupes de parenthèses"
            
            first_group = groups[0]
            second_group = groups[1]
            
            # Compter les cartes dans chaque groupe
            first_count = self.count_cards(first_group)
            second_count = self.count_cards(second_group)
            
            print(f"📊 Jeu #{game_number}: Groupe 1 = {first_count} cartes ({first_group}), Groupe 2 = {second_count} cartes ({second_group})")
            
            # Vérifier si chaque groupe a 3 cartes de couleurs différentes
            first_has_different_suits = (first_count == 3) and self.has_different_suits(first_group)
            second_has_different_suits = (second_count == 3) and self.has_different_suits(second_group)
            
            # NOUVELLE LOGIQUE DE DÉTERMINATION DU GAGNANT
            winner = None
            
            if first_has_different_suits and second_has_different_suits:
                # Les deux ont 3 cartes différentes → on ignore
                print(f"⚠️ Les deux groupes ont 3 cartes de couleurs différentes, message ignoré")
                return False, "Les deux groupes ont 3 couleurs différentes - pas d'enregistrement"
            elif first_has_different_suits and not second_has_different_suits:
                # Premier groupe a 3 cartes différentes → Victoire JOUEUR
                winner = 'Joueur'
                print(f"🎯 Premier groupe a 3 cartes différentes → Victoire JOUEUR")
            elif not first_has_different_suits and second_has_different_suits:
                # Deuxième groupe a 3 cartes différentes → Victoire BANQUIER
                winner = 'Banquier'
                print(f"🎯 Deuxième groupe a 3 cartes différentes → Victoire BANQUIER")
            else:
                # Aucun groupe n'a 3 cartes différentes → on ignore
                print(f"⚠️ Aucun groupe n'a 3 cartes de couleurs différentes, message ignoré")
                return False, "Aucun groupe avec 3 couleurs différentes"
            
            # Si on arrive ici, on a un gagnant valide
            
            # Extraire date et heure du message
            date_str, time_str = self.extract_datetime_from_message(message)
            
            # Créer l'entrée de résultat
            result_entry = {
                'numero': game_number,
                'date': date_str,
                'heure': time_str,
                'cartes_groupe1': first_group.strip(),
                'gagnant': winner,
                'message_complet': message[:200]  # Limiter la taille
            }
            
            # Ajouter et sauvegarder
            results.append(result_entry)
            self._save_yaml(results)
            
            print(f"✅ Résultat enregistré: Jeu #{game_number} - Gagnant: {winner} - {date_str} {time_str}")
            return True, f"Jeu #{game_number} enregistré - Gagnant: {winner}"
            
        except Exception as e:
            print(f"❌ Erreur traitement message: {e}")
            import traceback
            traceback.print_exc()
            return False, f"Erreur: {e}"
    
    def get_all_results(self) -> List[Dict[str, Any]]:
        """Récupère tous les résultats stockés"""
        return self._load_yaml()
    
    def get_stats(self) -> Dict[str, Any]:
        """Calcule les statistiques des résultats"""
        results = self._load_yaml()
        
        if not results:
            return {
                'total': 0,
                'joueur_victoires': 0,
                'banquier_victoires': 0,
                'taux_joueur': 0.0,
                'taux_banquier': 0.0
            }
        
        joueur_wins = sum(1 for r in results if r.get('gagnant') == 'Joueur')
        banquier_wins = sum(1 for r in results if r.get('gagnant') == 'Banquier')
        total = len(results)
        
        return {
            'total': total,
            'joueur_victoires': joueur_wins,
            'banquier_victoires': banquier_wins,
            'taux_joueur': (joueur_wins / total * 100) if total > 0 else 0.0,
            'taux_banquier': (banquier_wins / total * 100) if total > 0 else 0.0
        }
    
    def export_to_txt(self, file_path: str = None) -> Optional[str]:
        """Exporte tous les résultats en fichier Excel"""
        try:
            # Générer un nom de fichier avec date et heure si non fourni
            if file_path is None:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                file_path = f"resultats_{timestamp}.xlsx"
            
            results = self._load_yaml()
            
            # Créer un nouveau classeur Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Résultats"
            
            # Style pour l'en-tête
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # En-têtes de colonnes
            headers = ["Date & Heure", "Numéro", "Victoire (Joueur/Banquier)"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Largeur des colonnes
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 30
            
            if not results:
                # Si pas de résultats
                cell = ws.cell(row=2, column=1)
                cell.value = "Aucun résultat enregistré."
                cell.alignment = Alignment(horizontal="center")
            else:
                # Ajouter les données
                for row_num, result in enumerate(results, 2):
                    # Date et Heure
                    date_str = result.get('date', '')
                    heure_str = result.get('heure', '')
                    
                    if date_str and heure_str:
                        try:
                            date_parts = date_str.split('-')
                            if len(date_parts) == 3:
                                formatted_date = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]}"
                            else:
                                formatted_date = date_str
                        except:
                            formatted_date = date_str
                        
                        try:
                            heure_parts = heure_str.split(':')
                            if len(heure_parts) >= 2:
                                formatted_heure = f"{heure_parts[0]}:{heure_parts[1]}"
                            else:
                                formatted_heure = heure_str
                        except:
                            formatted_heure = heure_str
                        
                        date_heure = f"{formatted_date} - {formatted_heure}"
                    else:
                        date_heure = "N/A"
                    
                    # Numéro
                    numero = result.get('numero', 0)
                    numero_formatted = f"{numero:03d}"
                    
                    # Gagnant
                    gagnant = result.get('gagnant', 'N/A')
                    
                    # Écrire les données
                    cell_a = ws.cell(row=row_num, column=1)
                    cell_a.value = date_heure
                    cell_a.border = border
                    cell_a.alignment = Alignment(horizontal="left")
                    
                    cell_b = ws.cell(row=row_num, column=2)
                    cell_b.value = numero_formatted
                    cell_b.border = border
                    cell_b.alignment = Alignment(horizontal="center")
                    
                    cell_c = ws.cell(row=row_num, column=3)
                    cell_c.value = gagnant
                    cell_c.border = border
                    cell_c.alignment = Alignment(horizontal="center")
            
            # Sauvegarder le fichier
            wb.save(file_path)
            print(f"✅ Export Excel créé: {file_path}")
            return file_path
            
        except Exception as e:
            print(f"❌ Erreur export Excel: {e}")
            import traceback
            traceback.print_exc()
            return None
